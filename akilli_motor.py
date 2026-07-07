"""Akıllı Komisyon Motoru – 10 Modüllü Karar Sistemi.

Trendyol komisyon tarifelerini, sipariş geçmişini ve stok verilerini
birleştirerek her model+renk varyantı için optimal kademe önerisi üretir.

Saf analiz modülleri (mod1..mod10) akilli_motor_moduller.py'dedir; bu dosya
veri erişimi (DB satış/maliyet), orkestrasyon (run_full_analysis), Excel
giriş/çıkışı ve Flask route'larını içerir.

Modüller:
  1. Satış Hızı + Üretim Zekası
  2. Renk Segment Zekası
  3. Ürün Yaşam Döngüsü
  4. Portföy Matrisi (BCG)
  5. Gerçek Maliyet + Nakit Akış (iade beklenen-değerli)
  6. Senaryo Simülatörü (ölçülmüş elastikiyet)
  7. Karar Motoru (Skor Kartı, portföy-p90 normalizasyonu)
  8. Üretim Kapasite Tahsisi
  9. Sezon + Takvim Zekası (kategori öncelikli)
  10. Model Verimlilik Karşılaştırma
"""

from flask import (
    Blueprint, request, jsonify, render_template,
    send_file, flash, redirect, url_for,
)
import pandas as pd
import numpy as np
import os
import logging
from datetime import datetime, timedelta
from werkzeug.utils import secure_filename
from login_logout import login_required, roles_required

from akilli_motor_moduller import (
    STRATEGY_WEIGHTS, ELASTICITY_BASE,
    _resolve_commission_columns, _extract_model_from_sku,
    _extract_color_from_sku, _extract_color_from_tariff,
    beklenen_birim_kar, estimate_elasticity, hesapla_skor_olcek,
    mod1_satis_hizi, mod2_renk_segment, mod3_yasam_dongusu, mod4_portfoy,
    mod5_gercek_maliyet, mod6_senaryo, mod7_skor, mod8_kapasite,
    mod9_sezon, mod10_model_verimlilik,
)

logger = logging.getLogger(__name__)

akilli_motor_bp = Blueprint('akilli_motor', __name__)

UPLOAD_FOLDER = './uploads'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


# ═══════════════════════════════════════════════════════════════════════
#  YARDIMCI FONKSİYONLAR (dosya / Excel çıkışı)
# ═══════════════════════════════════════════════════════════════════════

def _allowed_file(filename: str) -> bool:
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def _write_tariff_output(src_path: str, out_path: str, header: list,
                         row_updates: dict, donem: str) -> int:
    """Trendyol komisyon tarife Excel'ini BİREBİR koruyarak çıktı üretir.

    openpyxl ile tüm workbook'u yeniden kaydetmek Trendyol'un kabul ettiği
    dosya yapısını bozuyordu (sharedStrings→inline string, calcChain kaybı,
    formül/stil farkları) → Trendyol reddediyordu. Oysa kullanıcı aynı dosyayı
    elle Excel'de doldurunca kabul ediliyor.

    Bu yüzden xlsx'i ZIP/XML seviyesinde açıp SADECE iki sütunun
    (YENİ TSF (FİYAT GÜNCELLE) + Tarife Seçimi) boş hücrelerini doldururuz;
    diğer her şey (stiller, formüller, paylaşılan metinler, tüm satır/sütunlar)
    byte-byte korunur. Değiştirilen hücre sayısını döndürür.
    """
    import re
    import zipfile
    from openpyxl.utils import get_column_letter

    def _col(name):
        return get_column_letter(header.index(name) + 1) if name in header else None

    yeni_letter = _col('YENİ TSF (FİYAT GÜNCELLE)')
    tarife_letter = _col('Tarife Seçimi')
    hes3_letter = _col('Hesaplanan Komisyon (3 Gün)')
    hes4_letter = _col('Hesaplanan Komisyon (4 Gün)')

    if not yeni_letter:
        import shutil
        shutil.copy(src_path, out_path)
        return 0

    tarife_value = f'{donem} Günlük Fiyat'

    def _esc(s: str) -> str:
        return s.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

    def _num(v) -> str:
        return '%g' % round(float(v), 2)

    # row_updates değerleri dict: {'yeni_tsf', 'hes3', 'hes4'} (geriye dönük:
    # düz sayı da kabul edilir).
    def _unpack(val):
        if isinstance(val, dict):
            return val.get('yeni_tsf'), val.get('hes3'), val.get('hes4')
        return val, None, None

    # Boş hücre güncellemeleri (YENİ TSF sayı, Tarife Seçimi metin)
    empty_updates: dict[str, tuple[str, str]] = {}
    # Formül hücresine cached değer enjeksiyonu (Hesaplanan Komisyon 3/4 Gün)
    formula_updates: dict[str, object] = {}  # ref → sayı (float) veya None ('-')
    for pos, val in row_updates.items():
        excel_row = int(pos) + 2  # 1. satır başlık
        yeni_tsf, hes3, hes4 = _unpack(val)
        empty_updates[f'{yeni_letter}{excel_row}'] = ('num', _num(yeni_tsf))
        if tarife_letter:
            empty_updates[f'{tarife_letter}{excel_row}'] = ('str', tarife_value)
        if hes3_letter:
            formula_updates[f'{hes3_letter}{excel_row}'] = hes3
        if hes4_letter:
            formula_updates[f'{hes4_letter}{excel_row}'] = hes4

    zin = zipfile.ZipFile(src_path)
    sheet_names = [n for n in zin.namelist()
                   if re.match(r'xl/worksheets/sheet1\.xml$', n)] or \
                  [n for n in zin.namelist()
                   if n.startswith('xl/worksheets/') and n.endswith('.xml')]
    sheet = sheet_names[0]
    xml = zin.read(sheet).decode('utf-8')

    changed = 0

    # 1) Boş hücreleri doldur (YENİ TSF + Tarife Seçimi)
    cell_re = re.compile(r'<c r="([A-Z]+\d+)"((?: [a-z]+="[^"]*")*)\s*(?:/>|></c>)')

    def _repl_empty(m):
        nonlocal changed
        ref = m.group(1)
        if ref not in empty_updates:
            return m.group(0)
        attrs = re.sub(r' t="[^"]*"', '', m.group(2))
        kind, val = empty_updates[ref]
        changed += 1
        if kind == 'num':
            return f'<c r="{ref}"{attrs}><v>{val}</v></c>'
        return f'<c r="{ref}"{attrs} t="inlineStr"><is><t>{_esc(val)}</t></is></c>'

    new_xml = cell_re.sub(_repl_empty, xml)

    # 2) Hesaplanan Komisyon formül hücrelerine cached değer ekle (formülü koru)
    fcell_re = re.compile(
        r'<c r="([A-Z]+\d+)"((?: [a-z]+="[^"]*")*)>(<f[^>]*>.*?</f>|<f[^>]*/>)(<v>.*?</v>)?</c>'
    )

    def _repl_formula(m):
        nonlocal changed
        ref = m.group(1)
        if ref not in formula_updates:
            return m.group(0)
        attrs = re.sub(r' t="[^"]*"', '', m.group(2))
        fpart = m.group(3)
        hv = formula_updates[ref]
        changed += 1
        if hv is None:
            return f'<c r="{ref}"{attrs} t="str">{fpart}<v>-</v></c>'
        return f'<c r="{ref}"{attrs}>{fpart}<v>{_num(hv)}</v></c>'

    new_xml = fcell_re.sub(_repl_formula, new_xml)

    with zipfile.ZipFile(out_path, 'w', zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = new_xml.encode('utf-8') if item.filename == sheet else zin.read(item.filename)
            zout.writestr(item, data)
    zin.close()
    return changed


# ═══════════════════════════════════════════════════════════════════════
#  VERİ ERİŞİMİ (DB)
# ═══════════════════════════════════════════════════════════════════════

def _query_sales_data(days: int = 90) -> pd.DataFrame:
    """Son N günlük sipariş verisini DB'den çeker (varsayılan 90)."""
    from models import OrderShipped, OrderDelivered

    cutoff = datetime.now() - timedelta(days=days)
    rows = []

    for model_cls in [OrderShipped, OrderDelivered]:
        try:
            orders = model_cls.query.filter(
                model_cls.order_date >= cutoff
            ).with_entities(
                model_cls.merchant_sku,
                model_cls.product_color,
                model_cls.amount,
                model_cls.commission,
                model_cls.order_date,
                model_cls.quantity,
            ).all()

            for o in orders:
                model_code = _extract_model_from_sku(o.merchant_sku)
                if not model_code:
                    continue
                rows.append({
                    'model_kodu': model_code,
                    'renk': o.product_color or 'Standart',
                    'tutar': float(o.amount or 0),
                    'komisyon': float(o.commission or 0),
                    'tarih': o.order_date,
                    'adet': int(o.quantity or 1),
                })
        except Exception as e:
            logger.warning(f"Sipariş verisi çekme hatası ({model_cls.__name__}): {e}")

    if not rows:
        return pd.DataFrame(columns=['model_kodu', 'renk', 'tutar', 'komisyon', 'tarih', 'adet'])

    df = pd.DataFrame(rows)
    df['tarih'] = pd.to_datetime(df['tarih'], errors='coerce')
    return df


def _model_cost_map_tl(model_codes: list[str]) -> dict[str, float]:
    """Tedarikçi sayfasındaki model maliyetlerini TL olarak döndürür.

    Kaynak agent_api model-cost API'siyle aynı: ModelDirekMaliyet (>0)
    öncelikli, yoksa ModelMaliyet kalem toplamı (USD) → güncel kurla TL.
    Model kodları baştaki sıfır farkına karşı iki varyantla da aranır
    ("0172" ↔ "172"). Kur/DB hatasında boş dict döner ve motor formdaki
    manuel maliyete düşer — analiz asla bu yüzden patlamaz.
    """
    try:
        from siparis_fisi import _usd_maliyet_map
        from profit import _fetch_usd_try

        rate = _fetch_usd_try()
        if not rate or rate <= 0:
            logger.warning("Model maliyeti: USD/TL kuru alınamadı, manuel maliyete düşülüyor")
            return {}

        arananlar = set()
        for m in model_codes:
            s = str(m).strip()
            if not s:
                continue
            arananlar.add(s)
            arananlar.add(s.lstrip('0') or '0')

        usd_map = _usd_maliyet_map(list(arananlar))
        sonuc = {}
        for m in model_codes:
            s = str(m).strip()
            usd = usd_map.get(s) or usd_map.get(s.lstrip('0') or '0') or 0
            if usd > 0:
                sonuc[s] = round(float(usd) * float(rate), 2)
        return sonuc
    except Exception as e:
        logger.warning(f"Model maliyet haritası çekilemedi: {e}")
        return {}


def _kaydet_oneri_gecmisi(results: list[dict], strateji: str,
                          tarife_donemi: str) -> int:
    """TUT dışındaki önerileri geri besleme için MotorOneriLog'a yazar.

    Aynı model+renk için günde bir kayıt (tekrar koşular spam üretmesin).
    Tablo henüz migrate edilmemişse sessizce 0 döner — analiz akışını bozmaz.
    """
    from models import db, MotorOneriLog

    try:
        gun_basi = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        bugunkuler = {
            (r.model_kodu, r.renk)
            for r in MotorOneriLog.query.filter(
                MotorOneriLog.created_at >= gun_basi
            ).with_entities(MotorOneriLog.model_kodu, MotorOneriLog.renk).all()
        }

        eklenen = 0
        for r in results:
            if r['aksiyon'] == 'TUT':
                continue
            if (r['model_kodu'], r['renk']) in bugunkuler:
                continue
            db.session.add(MotorOneriLog(
                model_kodu=r['model_kodu'],
                renk=r['renk'],
                eff_renk=r.get('eff_renk', r['renk']),
                aksiyon=r['aksiyon'],
                mevcut_fiyat=r['guncel_tsf'],
                onerilen_fiyat=r['onerilen_fiyat'],
                onerilen_kademe=r['onerilen_kademe'],
                skor=r['onerilen_skor'],
                gunluk_satis=r['gunluk_satis'],
                stok=r['stok'],
                strateji=strateji,
                tarife_donemi=str(tarife_donemi),
            ))
            eklenen += 1
        if eklenen:
            db.session.commit()
        return eklenen
    except Exception as e:
        try:
            db.session.rollback()
        except Exception:
            pass
        logger.warning(
            "Öneri geçmişi kaydedilemedi (tablo migrate edilmemiş olabilir: "
            f"scripts/create_motor_oneri_log_table.py): {e}"
        )
        return 0


# ═══════════════════════════════════════════════════════════════════════
#  ANA ANALİZ ORKESTRATÖRÜ
# ═══════════════════════════════════════════════════════════════════════

def run_full_analysis(tariff_df: pd.DataFrame, params: dict,
                      excluded_models: list[str] | None = None,
                      included_models: list[str] | None = None,
                      include_only: bool = False,
                      tarife_donemi: str = '3',
                      haftalik_kapasite: int = 0) -> dict:
    """Tüm modülleri çalıştırıp birleşik sonuç üretir.

    include_only=True ve included_models dolu ise SADECE bu modeller analiz
    edilip fiyat güncellemesi (indirim) alır; diğer tüm modeller dokunulmadan
    kalır (beyaz liste modu).

    tarife_donemi ('3' veya '4'): Yeni Trendyol formatında komisyon oranları
    hangi teslimat tarifesinden ('3 Gün' / '4 Gün') okunacağını belirler.

    haftalik_kapasite: Modül 8 için haftalık üretim kapasitesi (çift);
    0 = kapasite kısıtı yok.
    """

    # ── Veri Hazırlığı ──────────────────────────────────────────────
    tariff_df.columns = [str(c).strip() for c in tariff_df.columns]

    # Seçilen teslimat tarifesine göre komisyon kademesi sütunları.
    # '7 Günlük' için Excel'de ayrı komisyon bloğu yoktur (yalnızca 3/4 Gün);
    # komisyon 3 Gün setinden okunur ama çıktıya '7 Günlük Fiyat' yazılır.
    tarife_donemi = str(tarife_donemi).strip()
    if tarife_donemi not in ('3', '4', '7'):
        tarife_donemi = '3'
    kom_period = '4' if tarife_donemi == '4' else '3'
    kom_cols = _resolve_commission_columns(tariff_df.columns, kom_period)
    # Hesaplanan Komisyon (3/4 Gün) cached değerlerini üretmek için her iki
    # tarife setinin komisyon sütunlarını da sayısala çeviririz.
    kom_cols_3 = _resolve_commission_columns(tariff_df.columns, '3')
    kom_cols_4 = _resolve_commission_columns(tariff_df.columns, '4')

    limit_cols = [
        '1.Fiyat Alt Limit', '2.Fiyat Üst Limiti', '2.Fiyat Alt Limit',
        '3.Fiyat Üst Limiti', '3.Fiyat Alt Limit', '4.Fiyat Üst Limiti',
    ]
    numeric_cols = [
        *limit_cols,
        *kom_cols, *kom_cols_3, *kom_cols_4,
        'KOMİSYONA ESAS FİYAT', 'GÜNCEL KOMİSYON', 'GÜNCEL TSF',
    ]
    for col in numeric_cols:
        if col in tariff_df.columns:
            tariff_df[col] = pd.to_numeric(tariff_df[col], errors='coerce').fillna(0)

    tariff_df['_RENK'] = tariff_df.apply(
        lambda r: _extract_color_from_tariff(
            r.get('SATICI STOK KODU', ''), str(r.get('MODEL KODU', '')), r.get('BEDEN')
        ), axis=1,
    )

    # Hariç model setleri: hem ham hem de baştaki sıfırlar atılmış halini ekle
    # (Excel hücresi sayı saklıyorsa "0172" → 172, kullanıcı "0172" yazmış olabilir)
    excluded = set()
    for m in (excluded_models or []):
        s = m.strip().upper()
        if not s:
            continue
        excluded.add(s)
        excluded.add(s.lstrip('0') or '0')

    # Beyaz liste (yalnız bu modellere indirim uygula): excluded ile aynı normalize
    included = set()
    for m in (included_models or []):
        s = m.strip().upper()
        if not s:
            continue
        included.add(s)
        included.add(s.lstrip('0') or '0')
    include_only = include_only and bool(included)

    # Sipariş verisini çek
    sales_df = _query_sales_data()

    # Ölçülmüş fiyat elastikiyeti (yetersiz veri → ELASTICITY_BASE yedeği)
    olculen_elastikiyet, elastikiyet_gozlem = estimate_elasticity(sales_df)
    elastic_base = olculen_elastikiyet if olculen_elastikiyet is not None else ELASTICITY_BASE

    # Model bazlı gerçek maliyet (tedarikçi sayfası, USD→TL); yoksa form değeri
    model_codes = sorted({
        str(m).strip() for m in tariff_df.get('MODEL KODU', pd.Series(dtype=object)).dropna()
        if str(m).strip()
    })
    cost_map = _model_cost_map_tl(model_codes)

    uretim_suresi = params.get('uretim_suresi', 25)
    guvenlik_gun = params.get('guvenlik_gun', 5)
    strategy = params.get('strateji', 'dengeli')
    weights = STRATEGY_WEIGHTS.get(strategy, STRATEGY_WEIGHTS['dengeli'])

    # ── Renk Varyantı Bazında Gruplama ──────────────────────────────
    grouped = tariff_df.groupby(['MODEL KODU', '_RENK'], sort=False)

    results = []
    production_alerts = []

    # İlk geçiş: temel verileri topla (portföy ortalamaları + skor ölçeği için)
    pre_data = []
    for (model_kodu, renk), group in grouped:
        model_str = str(model_kodu).strip()
        model_upper = model_str.upper()
        if model_upper in excluded or (model_upper.lstrip('0') or '0') in excluded:
            continue
        # Beyaz liste modu: listede olmayan modeli atla (indirim uygulanmaz, dokunulmaz)
        if include_only and not (model_upper in included or (model_upper.lstrip('0') or '0') in included):
            continue
        first = group.iloc[0]
        stok = int(group['STOK'].sum()) if 'STOK' in group.columns else 0
        guncel_tsf = float(first.get('GÜNCEL TSF', 0) or 0)
        guncel_kom = float(first.get('GÜNCEL KOMİSYON', 0) or 0)

        # Model bazlı maliyet; DB'de yoksa formdaki manuel değer
        maliyet = cost_map.get(model_str, params.get('maliyet', 0))
        maliyet_kaynak = 'db' if model_str in cost_map else 'manuel'
        vparams = {**params, 'maliyet': maliyet}
        birim_kar = beklenen_birim_kar(guncel_tsf, guncel_kom, vparams)

        # Renk eşleştirme (pre-pass)
        eff_renk = renk
        mm = sales_df['model_kodu'] == model_str
        if mm.any():
            db_rs = sales_df[mm]['renk'].unique()
            if renk not in db_rs:
                rl = renk.lower()
                for dr in db_rs:
                    if rl in dr.lower() or dr.lower() in rl:
                        eff_renk = dr
                        break

        is_stok_eritme = strategy == 'stok_eritme'
        m1 = mod1_satis_hizi(sales_df, model_str, eff_renk, stok, uretim_suresi, guvenlik_gun, stok_eritme=is_stok_eritme)
        pre_data.append({
            'model': model_str, 'renk': renk, 'eff_renk': eff_renk, 'stok': stok,
            'gunluk_satis': m1['gunluk_satis'], 'birim_kar': birim_kar,
            'maliyet': maliyet, 'maliyet_kaynak': maliyet_kaynak,
            'group': group, 'first': first, 'm1': m1,
        })

    if not pre_data:
        return {'success': True, 'results': [], 'alerts': [], 'stats': {},
                'portfolio': {}, 'row_updates': {}, 'tarife_donemi': tarife_donemi,
                'kapasite_plan': None, 'model_verimlilik': []}

    ort_gunluk = np.mean([p['gunluk_satis'] for p in pre_data]) if pre_data else 0.01
    ort_birim_kar = np.mean([p['birim_kar'] for p in pre_data]) if pre_data else 0

    # Skor normalizasyonu: portföyün kendi p90 dağılımı (sabit eşik yerine)
    skor_olcek = hesapla_skor_olcek(pre_data)

    # ── İkinci Geçiş: Tüm Modülleri Çalıştır ───────────────────────
    for p in pre_data:
        model_str = p['model']
        renk = p['renk']
        group = p['group']
        first = p['first']
        stok = p['stok']
        m1 = p['m1']
        effective_renk = p['eff_renk']
        vparams = {**params, 'maliyet': p['maliyet']}

        guncel_tsf = float(first.get('GÜNCEL TSF', 0) or 0)
        guncel_kom = float(first.get('GÜNCEL KOMİSYON', 0) or 0)
        kategori = str(first.get('KATEGORİ', ''))

        k1 = float(first.get(kom_cols[0], 0) or 0)
        k2 = float(first.get(kom_cols[1], 0) or 0)
        k3 = float(first.get(kom_cols[2], 0) or 0)
        k4 = float(first.get(kom_cols[3], 0) or 0)
        ust2 = float(first.get('2.Fiyat Üst Limiti', 0) or 0)
        ust3 = float(first.get('3.Fiyat Üst Limiti', 0) or 0)
        ust4 = float(first.get('4.Fiyat Üst Limiti', 0) or 0)

        # Modül 2: Renk Segment
        m2 = mod2_renk_segment(sales_df, model_str, effective_renk)

        # Modül 3: Yaşam Döngüsü
        m3 = mod3_yasam_dongusu(sales_df, model_str, effective_renk)

        # Modül 4: Portföy (BCG)
        m4 = mod4_portfoy(m1['gunluk_satis'], p['birim_kar'], ort_gunluk, ort_birim_kar)

        # Modül 9: Sezon (kategori öncelikli, renk ikincil)
        m9 = mod9_sezon(renk, kategori)

        # ── Her Kademe İçin Senaryo + Skor ──────────────────────────
        tiers = []
        params_with_price = {**vparams, '_mevcut_fiyat': guncel_tsf}

        tier_configs = [
            (1, guncel_tsf, k1),
            (2, ust2, k2),
            (3, ust3, k3),
            (4, ust4, k4),
        ]

        for kademe, fiyat, kom in tier_configs:
            if fiyat <= 0 or kom <= 0:
                continue

            ort_stokta_gun = stok / m1['gunluk_satis'] / 2 if m1['gunluk_satis'] > 0 else 90
            m5 = mod5_gercek_maliyet(fiyat, kom, vparams, ort_stokta_gun)
            m6 = mod6_senaryo(fiyat, kom, m1['gunluk_satis'], stok,
                              m2['segment'], params_with_price, uretim_suresi,
                              elastic_base=elastic_base)
            skor = mod7_skor(m6, m1, m2, m3, m4, m5, weights, stok,
                             uretim_suresi, kademe, olcek=skor_olcek)

            tiers.append({
                'kademe': kademe,
                'fiyat': round(fiyat, 2),
                'komisyon_oran': kom,
                'net_kar': m5['net_kar'],
                'kar_marji': m5['kar_marji'],
                'tahmini_gunluk': m6['tahmini_gunluk_satis'],
                'tahmini_aylik': m6['tahmini_aylik_satis'],
                'aylik_kar': m6['aylik_kar'],
                'stok_bitis_gun': m6['stok_bitis_gun'],
                'roi_hizi': m5['roi_hizi'],
                'skor': skor,
                'uretim_uyari': m6['uretim_uyari'],
            })

        if not tiers:
            continue

        best = max(tiers, key=lambda t: t['skor'])
        mevcut = tiers[0]  # kademe 1

        # Aksiyon belirleme
        if best['kademe'] == 1:
            aksiyon = 'TUT'
        elif m3['asama'] == 'olu':
            aksiyon = 'TASFİYE'
        else:
            aksiyon = 'DÜŞÜR'

        # Üretim alarmı
        if m1['uretim_alarm'] == 'kritik':
            production_alerts.append({
                'model': model_str, 'renk': renk, 'stok': stok,
                'kalan_gun': m1['kalan_gun'],
                'gunluk_satis': m1['gunluk_satis'],
                'mesaj': f'{model_str}-{renk}: {m1["kalan_gun"]:.0f} gün stok, üretim {uretim_suresi} gün → HEMEN BAŞLAT',
            })

        results.append({
            'model_kodu': model_str,
            'renk': renk,
            'eff_renk': effective_renk,
            'urun_ismi': str(first.get('ÜRÜN İSMİ', '')),
            'kategori': kategori,
            'stok': stok,
            'varyant': len(group),
            'guncel_tsf': guncel_tsf,
            'guncel_komisyon': guncel_kom,
            'maliyet': round(p['maliyet'], 2),
            'maliyet_kaynak': p['maliyet_kaynak'],
            # Modül 1
            'gunluk_satis': m1['gunluk_satis'],
            'aylik_satis': m1['aylik_satis'],
            'kalan_gun': m1['kalan_gun'],
            'trend': m1['trend'],
            'trend_katsayi': m1['trend_katsayi'],
            'uretim_alarm': m1['uretim_alarm'],
            'uretim_emri_gun': m1['uretim_emri_gun'],
            # Modül 2
            'renk_segment': m2['segment'],
            'segment_aciklama': m2['aciklama'],
            'satis_orani': m2['satis_orani'],
            # Modül 3
            'yasam_asama': m3['asama'],
            'yasam_aciklama': m3['aciklama'],
            'momentum': m3['momentum'],
            # Modül 4
            'bcg_grup': m4['grup'],
            'bcg_icon': m4['icon'],
            'bcg_aksiyon': m4['aksiyon'],
            # Modül 9
            'sezon_uyum': m9['sezon_uyum'],
            'sezon_uyari': m9['uyari'],
            # Karar
            'aksiyon': aksiyon,
            'onerilen_kademe': best['kademe'],
            'onerilen_fiyat': best['fiyat'],
            'onerilen_komisyon': best['komisyon_oran'],
            'onerilen_skor': best['skor'],
            'mevcut_skor': mevcut['skor'],
            'skor_fark': round(best['skor'] - mevcut['skor'], 1),
            'onerilen_aylik_kar': best['aylik_kar'],
            'mevcut_aylik_kar': mevcut['aylik_kar'],
            'kar_fark_aylik': round(best['aylik_kar'] - mevcut['aylik_kar'], 0),
            # Tüm kademeler
            'kademeler': tiers,
        })

    # ── Portföy Dağılımı ────────────────────────────────────────────
    portfolio = {'yildiz': 0, 'nakit': 0, 'firsatci': 0, 'sorun': 0}
    for r in results:
        portfolio[r['bcg_grup']] = portfolio.get(r['bcg_grup'], 0) + 1

    # ── Modül 8: Üretim Kapasite Tahsisi ────────────────────────────
    kapasite_plan = mod8_kapasite(results, uretim_suresi, guvenlik_gun, haftalik_kapasite)

    # ── Modül 10: Model Verimlilik Karşılaştırma ────────────────────
    model_verimlilik = mod10_model_verimlilik(results)

    # ── Satır Bazında Güncelleme Haritası ───────────────────────────
    # Çıktı Excel'i, orijinal yüklenen dosya openpyxl ile açılıp SADECE iki
    # hücre (YENİ TSF + Tarife Seçimi) yazılarak üretilir. Bu yüzden burada
    # DataFrame'i yeniden yazmıyoruz; yalnızca "hangi satır → hangi yeni fiyat"
    # haritasını döndürüyoruz. Böylece formüller (Hesaplanan Komisyon), veri
    # tipleri (BARKOD/BEDEN/STOK) ve TÜM satırlar birebir korunur.
    # df konum indeksi i → Excel satırı i+2 (1. satır başlık).
    # Her güncellenen satır için Trendyol'un 'Hesaplanan Komisyon (3/4 Gün)'
    # formülünün vereceği değeri de hesaplarız (Excel'in cached value'su yerine).
    apply3 = tarife_donemi in ('3', '7')
    apply4 = tarife_donemi in ('4', '7')

    def _calc_hes(price, lims, koms):
        """Trendyol formül mantığı: fiyata göre uygun komisyon kademesi."""
        alt1, ust2, alt2, ust3, alt3, ust4 = lims
        k1, k2, k3, k4 = koms
        if price >= alt1:
            return k1 if k1 else None
        if alt2 <= price <= ust2:
            return k2 if k2 else None
        if alt3 <= price <= ust3:
            return k3 if k3 else None
        if price <= ust4:
            return k4 if k4 else None
        return None

    model_norm = tariff_df['MODEL KODU'].astype(str).str.strip()
    row_updates = {}
    for r in results:
        if r['aksiyon'] == 'TUT':
            continue
        mask = (model_norm == r['model_kodu']) & (tariff_df['_RENK'] == r['renk'])
        fiyat = round(float(r['onerilen_fiyat']), 2)
        for pos in np.where(mask.to_numpy())[0]:
            row = tariff_df.iloc[pos]
            lims = [float(row.get(c, 0) or 0) for c in limit_cols]
            hes3 = _calc_hes(fiyat, lims, [float(row.get(c, 0) or 0) for c in kom_cols_3]) if apply3 else None
            hes4 = _calc_hes(fiyat, lims, [float(row.get(c, 0) or 0) for c in kom_cols_4]) if apply4 else None
            row_updates[int(pos)] = {'yeni_tsf': fiyat, 'hes3': hes3, 'hes4': hes4}

    # ── İstatistikler ───────────────────────────────────────────────
    db_maliyetli_modeller = {r['model_kodu'] for r in results if r['maliyet_kaynak'] == 'db'}
    stats = {
        'toplam_varyant': len(results),
        'tut': sum(1 for r in results if r['aksiyon'] == 'TUT'),
        'dusur': sum(1 for r in results if r['aksiyon'] == 'DÜŞÜR'),
        'tasfiye': sum(1 for r in results if r['aksiyon'] == 'TASFİYE'),
        'uretim_alarm': len(production_alerts),
        'toplam_stok': sum(r['stok'] for r in results),
        'mevcut_aylik_kar': sum(r['mevcut_aylik_kar'] for r in results),
        'onerilen_aylik_kar': sum(r['onerilen_aylik_kar'] for r in results),
        'maliyet_db_model': len(db_maliyetli_modeller),
        'toplam_model': len({r['model_kodu'] for r in results}),
        'satis_kaydi': int(len(sales_df)),
        'elastikiyet': round(elastic_base, 2),
        'elastikiyet_kaynak': 'olcum' if olculen_elastikiyet is not None else 'varsayilan',
        'elastikiyet_gozlem': elastikiyet_gozlem,
    }
    stats['kar_artis'] = round(stats['onerilen_aylik_kar'] - stats['mevcut_aylik_kar'], 0)

    return {
        'success': True,
        'results': sorted(results, key=lambda x: x['skor_fark'], reverse=True),
        'alerts': sorted(production_alerts, key=lambda x: x['kalan_gun']),
        'stats': stats,
        'portfolio': portfolio,
        'row_updates': row_updates,
        'tarife_donemi': tarife_donemi,
        'kapasite_plan': kapasite_plan,
        'model_verimlilik': model_verimlilik,
    }


# ═══════════════════════════════════════════════════════════════════════
#  FLASK ROUTE'LARI
# ═══════════════════════════════════════════════════════════════════════

@akilli_motor_bp.route('/akilli-motor', methods=['GET'])
@login_required
@roles_required('admin')
def akilli_motor_sayfasi():
    return render_template('akilli_motor.html')


@akilli_motor_bp.route('/akilli-motor/analiz', methods=['POST'])
@login_required
@roles_required('admin')
def akilli_motor_analiz():
    if 'excel_file' not in request.files:
        return jsonify({'success': False, 'message': 'Dosya yüklenmedi'}), 400

    file = request.files['excel_file']
    if not file.filename or not _allowed_file(file.filename):
        return jsonify({'success': False, 'message': 'Geçersiz dosya'}), 400

    filename = secure_filename(file.filename)
    save_path = os.path.join(UPLOAD_FOLDER, f"motor_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}")

    try:
        file.save(save_path)
        ext = filename.rsplit('.', 1)[1].lower()
        df = pd.read_excel(save_path, engine='openpyxl' if ext == 'xlsx' else None)
        # MODEL KODU baştaki sıfırları korumak için string'e zorla (örn: "0172" → 172 olmasın)
        if 'MODEL KODU' in df.columns:
            def _normalize_model_kodu(v):
                if pd.isna(v):
                    return ''
                if isinstance(v, float) and v.is_integer():
                    return str(int(v))
                return str(v).strip()
            df['MODEL KODU'] = df['MODEL KODU'].apply(_normalize_model_kodu)

        params = {
            'maliyet': float(request.form.get('maliyet', 0) or 0),
            'kargo': float(request.form.get('kargo', 20) or 20),
            'iade_orani': float(request.form.get('iade_orani', 8) or 8),
            'paketleme': float(request.form.get('paketleme', 5) or 5),
            'sermaye_maliyeti': float(request.form.get('sermaye_maliyeti', 2.5) or 2.5),
            'uretim_suresi': int(request.form.get('uretim_suresi', 25) or 25),
            'guvenlik_gun': int(request.form.get('guvenlik_gun', 5) or 5),
            'strateji': request.form.get('strateji', 'dengeli'),
        }

        tarife_donemi = request.form.get('tarife_donemi', '3')
        haftalik_kapasite = int(float(request.form.get('haftalik_kapasite', 0) or 0))

        excluded_raw = request.form.get('excluded_models', '')
        excluded = [m.strip() for m in excluded_raw.replace('\n', ',').replace(';', ',').split(',') if m.strip()]

        included_raw = request.form.get('included_models', '')
        included = [m.strip() for m in included_raw.replace('\n', ',').replace(';', ',').split(',') if m.strip()]
        include_only = request.form.get('include_only', '') in ('1', 'true', 'True', 'on')

        analysis = run_full_analysis(df, params, excluded, included, include_only,
                                     tarife_donemi, haftalik_kapasite)

        if not analysis['success']:
            return jsonify(analysis), 400

        # Geri besleme: TUT dışı önerileri logla (başarısızlık analizi bozmaz).
        # Satış verisi hiç çekilemediyse LOGLAMA: veri kesintisinde her şey
        # 'olu'→TASFİYE görünür ve geçmiş çöp kayıtlarla kirlenir.
        if analysis.get('stats', {}).get('satis_kaydi', 0) > 0:
            analysis['oneri_kayit'] = _kaydet_oneri_gecmisi(
                analysis['results'], params['strateji'], analysis.get('tarife_donemi', tarife_donemi))
        else:
            analysis['oneri_kayit'] = 0

        # ── Çıktı: yüklenen dosyayı BİREBİR koru, sadece 2 sütunu yaz ──
        # ZIP/XML seviyesinde cerrahi düzenleme → formüller, stiller, paylaşılan
        # metinler ve tüm yapı korunur (openpyxl yeniden-kaydı Trendyol'u
        # reddettiriyordu). Bkz. _write_tariff_output.
        import openpyxl
        row_updates = analysis.get('row_updates', {})
        donem = analysis.get('tarife_donemi', tarife_donemi)

        wb_h = openpyxl.load_workbook(save_path)
        header = [c.value for c in wb_h.active[1]]
        wb_h.close()

        output_path = save_path.replace('.xls', '_motor_optimized.xls')
        _write_tariff_output(save_path, output_path, header, row_updates, donem)

        analysis.pop('row_updates', None)
        analysis['download_file'] = os.path.basename(output_path)
        analysis['guncellenen_satir'] = len(row_updates)

        return jsonify(analysis)

    except Exception as e:
        logger.exception("Akıllı motor analiz hatası")
        return jsonify({'success': False, 'message': f'Hata: {str(e)}'}), 500
    finally:
        try:
            if os.path.exists(save_path):
                os.remove(save_path)
        except OSError:
            pass


@akilli_motor_bp.route('/akilli-motor/oneri-gecmisi', methods=['GET'])
@login_required
@roles_required('admin')
def akilli_motor_oneri_gecmisi():
    """Geçmiş önerileri gerçekleşen satış hızıyla karşılaştırır.

    Her öneri için: öneri tarihinden önceki 30 günün satış hızı vs sonraki
    (en fazla 30 günün) satış hızı. 7 günden taze öneriler 'çok yeni' sayılır.
    """
    try:
        from models import MotorOneriLog
        kayitlar = (MotorOneriLog.query
                    .order_by(MotorOneriLog.created_at.desc())
                    .limit(300).all())
    except Exception as e:
        logger.warning(f"Öneri geçmişi okunamadı: {e}")
        return jsonify({
            'success': False,
            'message': 'Öneri geçmişi tablosu bulunamadı. Sunucuda '
                       'scripts/create_motor_oneri_log_table.py çalıştırılmalı.',
        }), 200

    if not kayitlar:
        return jsonify({'success': True, 'oneriler': []})

    sales = _query_sales_data(days=180)
    now = datetime.now()

    oneriler = []
    for k in kayitlar:
        item = {
            'tarih': k.created_at.strftime('%d.%m.%Y'),
            'model_kodu': k.model_kodu,
            'renk': k.renk,
            'aksiyon': k.aksiyon,
            'mevcut_fiyat': k.mevcut_fiyat,
            'onerilen_fiyat': k.onerilen_fiyat,
            'onerilen_kademe': k.onerilen_kademe,
            'skor': k.skor,
            'strateji': k.strateji,
            'onceki_hiz': None, 'sonraki_hiz': None,
            'degisim_pct': None, 'durum': 'cok_yeni',
        }

        gecen_gun = (now - k.created_at).days
        if gecen_gun >= 7 and not sales.empty:
            eslesme_renk = k.eff_renk or k.renk
            mask = (sales['model_kodu'] == k.model_kodu) & (sales['renk'] == eslesme_renk)
            once_bas = k.created_at - timedelta(days=30)
            once_adet = sales[mask & (sales['tarih'] >= once_bas) & (sales['tarih'] < k.created_at)]['adet'].sum()
            onceki_hiz = once_adet / 30

            sonra_gun = min(gecen_gun, 30)
            sonra_bit = k.created_at + timedelta(days=sonra_gun)
            sonra_adet = sales[mask & (sales['tarih'] >= k.created_at) & (sales['tarih'] < sonra_bit)]['adet'].sum()
            sonraki_hiz = sonra_adet / max(sonra_gun, 1)

            item['onceki_hiz'] = round(onceki_hiz, 2)
            item['sonraki_hiz'] = round(sonraki_hiz, 2)
            if onceki_hiz > 0:
                item['degisim_pct'] = round((sonraki_hiz - onceki_hiz) / onceki_hiz * 100)
                if sonraki_hiz > onceki_hiz * 1.1:
                    item['durum'] = 'artti'
                elif sonraki_hiz < onceki_hiz * 0.9:
                    item['durum'] = 'dustu'
                else:
                    item['durum'] = 'ayni'
            else:
                item['durum'] = 'olculemedi'

        oneriler.append(item)

    return jsonify({'success': True, 'oneriler': oneriler})


@akilli_motor_bp.route('/akilli-motor/indir/<filename>', methods=['GET'])
@login_required
@roles_required('admin')
def akilli_motor_indir(filename: str):
    safe_name = secure_filename(filename)
    file_path = os.path.join(UPLOAD_FOLDER, safe_name)
    if not os.path.exists(file_path):
        flash('Dosya bulunamadı.', 'danger')
        return redirect(url_for('akilli_motor.akilli_motor_sayfasi'))

    return send_file(
        file_path, as_attachment=True,
        download_name=f"akilli_motor_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
