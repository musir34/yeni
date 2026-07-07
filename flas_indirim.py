"""Flaş İndirim Motoru — Trendyol flaş teklif Excel'i analiz + doldurma.

Akış (akıllı komisyon motoruyla aynı desen):
  1. Kullanıcı Trendyol'un flaş indirim teklif Excel'ini yükler
     ('TeklifÜrünleri' sayfası; her satır ürün × tarih penceresi).
  2. Motor her teklif için taban fiyat (maliyet + %iade beklenen değeri +
     komisyon + kargo/paketleme + min kâr) hesaplar, Trendyol'un önerdiği
     flaş fiyatla karşılaştırır → KATIL / KATILMA kararı ve satış/kâr tahmini.
  3. Çıktı Excel'i yüklenen dosyayı BİREBİR koruyarak üretilir; yalnızca
     'Senin Belirlediğin Flaş Fiyatı' sütununun boş hücreleri doldurulur
     (ZIP/XML cerrahi düzenleme — akilli_motor._write_tariff_output deseni).

Saf karar mantığı flas_indirim_moduller.py'dedir; model maliyeti ve satış
verisi erişimi akilli_motor'daki mevcut yardımcılarla paylaşılır.
"""

from flask import (
    Blueprint, request, jsonify, render_template,
    send_file, flash, redirect, url_for,
)
import pandas as pd
import os
import logging
from datetime import datetime
from werkzeug.utils import secure_filename
from login_logout import login_required, roles_required

from akilli_motor import _query_sales_data, _model_cost_map_tl, _allowed_file
from akilli_motor_moduller import ELASTICITY_BASE, estimate_elasticity
from flas_indirim_moduller import (
    GEREKLI_SUTUNLAR, FLAS_GORUNURLUK_VARSAYILAN,
    flas_karar, model_satis_ozeti, model_filtre_setleri, model_filtrede,
)

logger = logging.getLogger(__name__)

flas_indirim_bp = Blueprint('flas_indirim', __name__)

UPLOAD_FOLDER = './uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

FLAS_FIYAT_SUTUNU = 'Senin Belirlediğin Flaş Fiyatı'


def _write_flas_output(src_path: str, out_path: str, header: list,
                       row_updates: dict) -> int:
    """Flaş Excel'ini BİREBİR koruyarak yalnızca flaş fiyat sütununu doldurur.

    Trendyol'un dosya yapısı (stiller, inline string'ler, tüm satırlar)
    byte-byte korunur; sadece 'Senin Belirlediğin Flaş Fiyatı' sütununun
    BOŞ hücrelerine sayısal değer yazılır. Değiştirilen hücre sayısını döndürür.
    row_updates: {df_pozisyon_indeksi: fiyat} — Excel satırı = pozisyon + 2.
    """
    import re
    import zipfile
    from openpyxl.utils import get_column_letter

    if FLAS_FIYAT_SUTUNU not in header:
        import shutil
        shutil.copy(src_path, out_path)
        return 0
    kolon = get_column_letter(header.index(FLAS_FIYAT_SUTUNU) + 1)

    updates = {
        f'{kolon}{int(pos) + 2}': '%g' % round(float(fiyat), 2)
        for pos, fiyat in row_updates.items()
    }

    zin = zipfile.ZipFile(src_path)
    sheet_names = [n for n in zin.namelist()
                   if re.match(r'xl/worksheets/sheet1\.xml$', n)] or \
                  [n for n in zin.namelist()
                   if n.startswith('xl/worksheets/') and n.endswith('.xml')]
    sheet = sheet_names[0]
    xml = zin.read(sheet).decode('utf-8')

    changed = 0
    cell_re = re.compile(r'<c r="([A-Z]+\d+)"((?: [a-z]+="[^"]*")*)\s*(?:/>|></c>)')

    def _repl(m):
        nonlocal changed
        ref = m.group(1)
        if ref not in updates:
            return m.group(0)
        attrs = re.sub(r' t="[^"]*"', '', m.group(2))
        changed += 1
        return f'<c r="{ref}"{attrs} t="n"><v>{updates[ref]}</v></c>'

    new_xml = cell_re.sub(_repl, xml)

    with zipfile.ZipFile(out_path, 'w', zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = new_xml.encode('utf-8') if item.filename == sheet else zin.read(item.filename)
            zout.writestr(item, data)
    zin.close()
    return changed


def _f(v, default=0.0):
    """NaN-güvenli float dönüşümü (jsonify NaN üretmesin)."""
    try:
        f = float(v)
        return default if pd.isna(f) else f
    except (TypeError, ValueError):
        return default


def _s(v):
    return '' if v is None or (isinstance(v, float) and pd.isna(v)) else str(v)


def run_flas_analysis(df: pd.DataFrame, params: dict,
                      excluded_models: list[str] | None = None,
                      included_models: list[str] | None = None,
                      include_only: bool = False,
                      min_kar: float = 0.0,
                      gorunurluk: float = FLAS_GORUNURLUK_VARSAYILAN,
                      taban_asiminda: str = 'katilma') -> dict:
    """Tüm teklif satırlarını analiz eder; sonuç + satır güncelleme haritası."""
    df.columns = [str(c).strip() for c in df.columns]

    eksik = [c for c in GEREKLI_SUTUNLAR if c not in df.columns]
    if eksik:
        return {'success': False,
                'message': 'Bu dosya Trendyol flaş indirim teklif Excel\'ine '
                           f'benzemiyor — eksik sütunlar: {", ".join(eksik)}'}

    excluded = model_filtre_setleri(excluded_models)
    included = model_filtre_setleri(included_models)
    include_only = include_only and bool(included)

    sales_df = _query_sales_data()
    olculen, gozlem = estimate_elasticity(sales_df)
    elastic_base = olculen if olculen is not None else ELASTICITY_BASE

    model_codes = sorted({str(m).strip() for m in df['Model Kodu'].dropna() if str(m).strip()})
    cost_map = _model_cost_map_tl(model_codes)

    satis_cache: dict[str, dict] = {}
    results = []
    row_updates = {}
    atlanan_model = 0

    for pos, row in df.iterrows():
        model = str(row.get('Model Kodu', '') or '').strip()
        if not model:
            continue
        if model_filtrede(model, excluded):
            atlanan_model += 1
            continue
        if include_only and not model_filtrede(model, included):
            atlanan_model += 1
            continue

        # 24 saat / 3 saat penceresi: hangi fiyat sütunu doluysa o teklif
        t24 = _f(row.get('24 Saat Fiyat'))
        t3 = _f(row.get('3 Saat Fiyat'))
        if t24 > 0:
            pencere, pencere_saat, trendyol_fiyat = '24s', 24, t24
            baslangic = _s(row.get('24 Saat Flaş Başlangıç Tarihi'))
        elif t3 > 0:
            pencere, pencere_saat, trendyol_fiyat = '3s', 3, t3
            baslangic = _s(row.get('3 Saat Flaş Başlangıç Tarihi'))
        else:
            pencere, pencere_saat, trendyol_fiyat = '-', 24, 0.0
            baslangic = ''

        maliyet = cost_map.get(model, params.get('maliyet', 0))
        maliyet_kaynak = 'db' if model in cost_map else 'manuel'
        vparams = {**params, 'maliyet': maliyet}

        if model not in satis_cache:
            satis_cache[model] = model_satis_ozeti(sales_df, model)
        satis = satis_cache[model]

        offer = {
            'model_kodu': model,
            'stok': int(_f(row.get('Stok'))),
            'mevcut_fiyat': _f(row.get('Mevcut Fiyat')),
            'musteri_fiyat': _f(row.get('Müşterinin Gördüğü Fiyat')),
            'komisyon': _f(row.get('Mevcut Komisyon')),
            'trendyol_fiyat': trendyol_fiyat,
            'pencere_saat': pencere_saat,
        }
        karar = flas_karar(offer, satis, vparams, elastic_base,
                           min_kar=min_kar, gorunurluk=gorunurluk,
                           taban_asiminda=taban_asiminda)

        if karar['aksiyon'] == 'KATIL' and karar['oneri_fiyat']:
            row_updates[int(pos)] = karar['oneri_fiyat']

        results.append({
            'pos': int(pos),
            'model_kodu': model,
            'urun_adi': _s(row.get('Ürün Adı')),
            'kategori': _s(row.get('Kategori')),
            'pencere': pencere,
            'baslangic': baslangic,
            'gun': baslangic.split(' ')[0] if baslangic else '',
            'stok': offer['stok'],
            'mevcut_fiyat': offer['mevcut_fiyat'],
            'musteri_fiyat': offer['musteri_fiyat'],
            'komisyon': offer['komisyon'],
            'trendyol_fiyat': trendyol_fiyat,
            'maliyet': round(maliyet, 2),
            'maliyet_kaynak': maliyet_kaynak,
            'gunluk_satis': satis['gunluk_satis'],
            'trend': satis['trend'],
            **karar,
        })

    katilanlar = [r for r in results if r['aksiyon'] == 'KATIL']
    db_modeller = {r['model_kodu'] for r in results if r['maliyet_kaynak'] == 'db'}
    stats = {
        'toplam_teklif': len(results),
        'katil': len(katilanlar),
        'katilma': len(results) - len(katilanlar),
        'gun_sayisi': len({r['gun'] for r in results if r['gun']}),
        'urun_sayisi': len({(r['model_kodu'], r['urun_adi']) for r in results}),
        'tahmini_toplam_satis': round(sum(r['tahmini_satis'] for r in katilanlar), 0),
        'tahmini_toplam_kar': round(sum(r['tahmini_kar'] for r in katilanlar), 0),
        'stok_uyari': sum(1 for r in katilanlar if r['stok_uyari']),
        'atlanan_satir': atlanan_model,
        'maliyet_db_model': len(db_modeller),
        'toplam_model': len({r['model_kodu'] for r in results}),
        'elastikiyet': round(elastic_base, 2),
        'elastikiyet_kaynak': 'olcum' if olculen is not None else 'varsayilan',
        'elastikiyet_gozlem': gozlem,
        'satis_kaydi': int(len(sales_df)),
    }

    # KATILMA'ları öne al (kullanıcı önce nelerin dışarıda kaldığını görsün
    # istemez — tersine: en kârlı KATIL'lar önce, sonra KATILMA'lar)
    results.sort(key=lambda r: (r['aksiyon'] != 'KATIL', -(r['tahmini_kar'] or 0)))

    return {'success': True, 'results': results, 'stats': stats,
            'row_updates': row_updates}


# ═══════════════════════════════════════════════════════════════════════
#  FLASK ROUTE'LARI
# ═══════════════════════════════════════════════════════════════════════

@flas_indirim_bp.route('/flas-indirim', methods=['GET'])
@login_required
@roles_required('admin')
def flas_indirim_sayfasi():
    return render_template('flas_indirim.html')


@flas_indirim_bp.route('/flas-indirim/analiz', methods=['POST'])
@login_required
@roles_required('admin')
def flas_indirim_analiz():
    if 'excel_file' not in request.files:
        return jsonify({'success': False, 'message': 'Dosya yüklenmedi'}), 400

    file = request.files['excel_file']
    if not file.filename or not _allowed_file(file.filename):
        return jsonify({'success': False, 'message': 'Geçersiz dosya (xlsx/xls olmalı)'}), 400

    filename = secure_filename(file.filename)
    save_path = os.path.join(UPLOAD_FOLDER, f"flas_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}")

    try:
        file.save(save_path)
        df = pd.read_excel(save_path, dtype={'Model Kodu': str})

        params = {
            'maliyet': float(request.form.get('maliyet', 0) or 0),
            'kargo': float(request.form.get('kargo', 20) or 20),
            'iade_orani': float(request.form.get('iade_orani', 20) or 20),
            'paketleme': float(request.form.get('paketleme', 5) or 5),
        }
        min_kar = float(request.form.get('min_kar', 0) or 0)
        gorunurluk = float(request.form.get('gorunurluk', FLAS_GORUNURLUK_VARSAYILAN)
                           or FLAS_GORUNURLUK_VARSAYILAN)
        taban_asiminda = request.form.get('taban_asiminda', 'katilma')
        if taban_asiminda not in ('katilma', 'taban_yaz'):
            taban_asiminda = 'katilma'

        excluded_raw = request.form.get('excluded_models', '')
        excluded = [m.strip() for m in excluded_raw.replace('\n', ',').replace(';', ',').split(',') if m.strip()]
        included_raw = request.form.get('included_models', '')
        included = [m.strip() for m in included_raw.replace('\n', ',').replace(';', ',').split(',') if m.strip()]
        include_only = request.form.get('include_only', '') in ('1', 'true', 'True', 'on')

        analysis = run_flas_analysis(df, params, excluded, included, include_only,
                                     min_kar, gorunurluk, taban_asiminda)
        if not analysis['success']:
            return jsonify(analysis), 400

        # Çıktı: dosyayı birebir koru, sadece flaş fiyat sütununu doldur
        import openpyxl
        wb_h = openpyxl.load_workbook(save_path, read_only=True)
        header = [c.value for c in next(wb_h.active.iter_rows(min_row=1, max_row=1))]
        wb_h.close()

        row_updates = analysis.pop('row_updates', {})
        output_path = save_path.replace('.xls', '_flas_optimized.xls')
        yazilan = _write_flas_output(save_path, output_path, header, row_updates)

        analysis['download_file'] = os.path.basename(output_path)
        analysis['guncellenen_satir'] = yazilan

        return jsonify(analysis)

    except Exception as e:
        logger.exception("Flaş indirim analiz hatası")
        return jsonify({'success': False, 'message': f'Hata: {str(e)}'}), 500
    finally:
        try:
            if os.path.exists(save_path):
                os.remove(save_path)
        except OSError:
            pass


@flas_indirim_bp.route('/flas-indirim/indir/<filename>', methods=['GET'])
@login_required
@roles_required('admin')
def flas_indirim_indir(filename: str):
    safe_name = secure_filename(filename)
    file_path = os.path.join(UPLOAD_FOLDER, safe_name)
    if not os.path.exists(file_path):
        flash('Dosya bulunamadı.', 'danger')
        return redirect(url_for('flas_indirim.flas_indirim_sayfasi'))

    return send_file(
        file_path, as_attachment=True,
        download_name=f"flas_indirim_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
