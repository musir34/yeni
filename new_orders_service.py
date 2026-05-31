# new_orders_service.py
from flask import Blueprint, render_template, request, jsonify, redirect, url_for, send_file
from models import db, OrderHazirlaniyor, RafUrun, Product, Archive, BarcodeAlias
from barcode_alias_helper import normalize_barcode, strip_turkish
from weather_service import get_istanbul_time
from sqlalchemy.orm import load_only
import json
import traceback
import io
from datetime import datetime
from zoneinfo import ZoneInfo
from qr_utils import qr_utils_bp  # routes/__init__.py bu modülden export ediyor

new_orders_service_bp = Blueprint('new_orders_service', __name__)

IST = ZoneInfo("Europe/Istanbul")

# Kargo cut-off'una bu saatten az kalan siparişler "acil" kovasına düşer.
# Pazaryeri/kargo bazında ileride parametrik yapılabilir.
KARGO_ACIL_ESIK_SAAT = 24


def _to_ist(dt):
    """Naive ise IST varsay, aware ise IST'ye çevir."""
    if dt is None:
        return None
    if dt.tzinfo is None:
        return dt.replace(tzinfo=IST)
    return dt.astimezone(IST)


def _remaining_seconds(agreed_delivery_date):
    """agreed_delivery_date'e kalan saniye. Tarih yoksa None (sıralamada en sona)."""
    dd = _to_ist(agreed_delivery_date)
    if dd is None:
        return None
    return (dd - get_istanbul_time()).total_seconds()


def _format_remaining(seconds):
    """Kalan saniyeyi 'X gün Y saat Z dakika' metnine çevirir."""
    if seconds is None:
        return "Kalan Süre Yok"
    if seconds <= 0:
        return "0 dakika"
    total = int(seconds)
    days, rem = divmod(total, 86400)
    hours, rem = divmod(rem, 3600)
    minutes = rem // 60
    return f"{days} gün {hours} saat {minutes} dakika"


def _prep_model():
    """Hazırlama/toplama kuyruğu = Hazırlanıyor (stoğu teyit edilmiş) siparişler."""
    return OrderHazirlaniyor


def _parse_details(order):
    """details alanını güvenli şekilde liste olarak döndürür."""
    if not order.details:
        return []
    try:
        items = json.loads(order.details) if isinstance(order.details, str) else order.details
        return items if isinstance(items, list) else []
    except (json.JSONDecodeError, TypeError):
        return []


def _build_shelf_groups():
    """
    Hazırlanıyor (OrderHazirlaniyor — stoğu teyit edilmiş) TÜM tek ürünlü siparişleri raf bazında gruplar.
    Dönüş: [(raf_kodu, [{'order_number', 'model_code', 'color', 'size', 'barcode'}, ...]), ...]
            Alfabetik raf sırasına göre.
    """
    PrepModel = _prep_model()
    new_orders = PrepModel.query.order_by(PrepModel.order_date.asc()).all()

    # Önce tüm barkodları toplayıp Product'ları tek sorguda çekelim
    all_barcodes = set()
    order_items = []  # [(order, item_dict)]

    for order in new_orders:
        details = _parse_details(order)
        # Sadece TEK ürünlü siparişler
        if len(details) != 1:
            continue
        item = details[0]
        barcode = normalize_barcode(item.get('barcode', ''))
        if barcode:
            all_barcodes.add(barcode)
        order_items.append((order, item, barcode))

    # Ürün bilgilerini toplu çek
    products_dict = {}
    if all_barcodes:
        products = Product.query.filter(Product.barcode.in_(list(all_barcodes))).all()
        products_dict = {p.barcode: p for p in products}

    # Raf bilgilerini toplu çek — barkod başına TÜM rafları kalan adetleriyle tut.
    # Her "Toplu Hazırla" çağrısında sıfırdan tahsis yapacağız: stale atanan_raf'ları
    # düzeltir ve bir rafta tek ürün varken iki siparişe atanmasını engeller.
    barcode_rafs: dict[str, list[list]] = {}  # barcode -> [[raf_kodu, kalan_adet], ...]
    if all_barcodes:
        rafs = (RafUrun.query
                .filter(RafUrun.urun_barkodu.in_(list(all_barcodes)))
                .filter(RafUrun.adet > 0)
                .order_by(RafUrun.raf_kodu.asc(), RafUrun.adet.desc())
                .with_for_update(read=True)
                .all())
        for raf in rafs:
            barcode_rafs.setdefault(raf.urun_barkodu, []).append([raf.raf_kodu, raf.adet])

    # Grup oluştur
    shelf_groups: dict[str, list[dict]] = {}

    # Her siparişi sıfırdan yeniden tahsis et — eski atanan_raf değerine güvenmiyoruz.
    # Sipariş tarihine göre sıralı geldiği için ilk gelen ilk rezerv eder.
    for order, item, barcode in order_items:
        product = products_dict.get(barcode)
        model_code = (product.product_main_id if product and product.product_main_id
                      else item.get('sku', 'N/A'))
        color = product.color if product and product.color else item.get('color', 'N/A')
        size = product.size if product and product.size else item.get('size', 'N/A')

        # Bu barkodun henüz boşalmamış ilk rafını seç ve 1 adet rezerve et.
        raf_kodu: str | None = None
        for raf_entry in barcode_rafs.get(barcode, []):
            if raf_entry[1] > 0:
                raf_kodu = raf_entry[0]
                raf_entry[1] -= 1
                break

        stok_yetersiz = raf_kodu is None
        yeni_atanan_raf = None if stok_yetersiz else raf_kodu

        # Stale atamayı her zaman geçersiz kıl — eski değer ne olursa olsun yenisini yaz.
        if order.atanan_raf != yeni_atanan_raf:
            order.atanan_raf = yeni_atanan_raf

        display_raf = raf_kodu if raf_kodu else 'RAF YOK'
        shelf_groups.setdefault(display_raf, []).append({
            'order_number': order.order_number,
            'model_code': model_code,
            'color': color,
            'size': size,
            'barcode': barcode,
            'stok_yetersiz': stok_yetersiz,
            'customer_name': ' '.join(filter(None, [order.customer_name, order.customer_surname])),
        })

    db.session.commit()

    # Alfabetik raf sırası
    return sorted(shelf_groups.items(), key=lambda x: x[0])


@new_orders_service_bp.route('/prepare-new-orders', methods=['GET'])
def prepare_new_orders():
    """
    Yeni statüsündeki tek ürünlü siparişleri raf bazında gruplar ve ekranda gösterir.
    """
    try:
        sorted_shelves = _build_shelf_groups()
        total_orders = sum(len(items) for _, items in sorted_shelves)
        return render_template(
            'bulk_order_prepare.html',
            sorted_shelves=sorted_shelves,
            total_orders=total_orders,
        )
    except Exception as e:
        traceback.print_exc()
        return f"Bir hata oluştu: {e}", 500


@new_orders_service_bp.route('/prepare-new-orders/excel', methods=['GET'])
def prepare_new_orders_excel():
    """
    Yeni statüsündeki tek ürünlü siparişleri xlsx olarak indirir.
    Sütunlar: Model Kodu | Renk | Beden | Sipariş No
    Raf sırasına (alfabetik) göre sıralanmış.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        sorted_shelves = _build_shelf_groups()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Raf Toplama Listesi"

        # Başlık satırı (A1'den itibaren, boş satır/sütun yok)
        headers = ["Model Kodu", "Renk", "Beden", "Sipariş No"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="B76E79", end_color="B76E79", fill_type="solid")

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row_num = 2
        for raf_kodu, items in sorted_shelves:
            for item in items:
                ws.cell(row=row_num, column=1, value=item['model_code'])
                ws.cell(row=row_num, column=2, value=item['color'])
                ws.cell(row=row_num, column=3, value=item['size'])
                ws.cell(row=row_num, column=4, value=item['order_number'])
                row_num += 1
                if row_num > 2001:  # max 2000 veri satırı
                    break
            if row_num > 2001:
                break

        # Sütun genişliklerini otomatik ayarla
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max(12, max_len + 2)

        # Belleğe kaydet
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"raf_toplama_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
    except Exception as e:
        traceback.print_exc()
        return f"Excel oluşturulurken hata oluştu: {e}", 500


# ─────────────────────────────────────────────────────────────
#  A4 ETİKET YAZDIR
# ─────────────────────────────────────────────────────────────

@new_orders_service_bp.route('/prepare-new-orders/labels-print', methods=['GET'])
def prepare_new_orders_labels_print():
    """
    Yeni siparişler için A4 etiket baskı sayfası.
    Her etiket: sipariş no barkodu + raf kodu.
    """
    try:
        from barcode_utils import generate_barcode_data_uri
        sorted_shelves = _build_shelf_groups()

        labels = []
        for raf_kodu, items in sorted_shelves:
            for item in items:
                labels.append({
                    'order_number': item['order_number'],
                    'raf_kodu': raf_kodu,
                    'model_code': item['model_code'],
                    'color': item['color'],
                    'size': item['size'],
                    'customer_name': item.get('customer_name', ''),
                    'barcode_img': generate_barcode_data_uri(item['order_number'], show_text=False),
                })

        return render_template('order_labels_print.html', labels=labels)
    except Exception as e:
        traceback.print_exc()
        return f"Hata: {e}", 500


# ─────────────────────────────────────────────────────────────
#  BARKOD OKUT → EN ACİL SİPARİŞİ GETİR
# ─────────────────────────────────────────────────────────────

@new_orders_service_bp.route('/prepare-new-orders/scan', methods=['POST'])
def scan_barcode_to_order():
    """
    Okutulan ürün barkoduna ait, 'Hazırlanıyor' (OrderHazirlaniyor — stoğu teyit edilmiş) TEK ürünlü
    siparişlerden en acilini bulur. Etiket basmaya gerek kalmaz; picker ürünü
    raftan alıp barkodunu okutur, sistem o adedi en acil siparişe atar.

    Sıralama (iki kovalı):
      1) Kargoya KARGO_ACIL_ESIK_SAAT saatten az kalanlar → en az süre kalan önce
         (cut-off'u geçmiş/gecikmiş olanlar en başta).
      2) Diğerleri → standart: cut-off tarihine göre, tarihi olmayan en sona,
         eşitlikte en eski sipariş (FIFO) önce.

    Dönüş (JSON):
      { "found": true, "order_number": ..., "redirect_url": "/hazirla?...",
        "model_code", "color", "size", "barcode", "raf", "customer_name",
        "remaining_time", "urgent" }
      veya { "found": false, "message": ... }
    """
    try:
        payload = request.get_json(silent=True) or request.form
        raw_barcode = (payload.get('barcode') or '').strip()
        if not raw_barcode:
            return jsonify({"found": False, "message": "Barkod boş."}), 400

        # Scanned barkodu BİR KEZ normalize et (alias/Product sorguları burada biter).
        scanned = normalize_barcode(raw_barcode)
        if not scanned:
            return jsonify({"found": False, "message": "Geçersiz barkod."}), 400

        # scanned'e eşlenen TÜM ham barkod biçimlerini önceden çıkar (tek alias sorgusu).
        # Böylece sipariş döngüsünde normalize_barcode() çağırıp DB'ye gitmeyiz.
        accepted_exact = {scanned}
        for a in BarcodeAlias.query.filter_by(main_barcode=scanned).all():
            accepted_exact.add(a.alias_barcode)
        accepted_ascii = {strip_turkish(x).lower() for x in accepted_exact}

        def _matches(raw):
            if not raw:
                return False
            rb = str(raw).strip().replace(' ', '')
            if rb in accepted_exact:
                return True
            return strip_turkish(rb).lower() in accepted_ascii

        # Arşivdeki siparişleri ele (onlar /hazirla ekranında zaten açılmaz).
        archived_numbers = {r[0] for r in db.session.query(Archive.order_number).all()}

        threshold_seconds = KARGO_ACIL_ESIK_SAAT * 3600

        # Sadece gereken kolonları çek — uzak DB'den tam satır transferini azalt.
        PrepModel = _prep_model()
        orders = (PrepModel.query
                  .options(load_only(
                      PrepModel.order_number,
                      PrepModel.details,
                      PrepModel.agreed_delivery_date,
                      PrepModel.order_date,
                      PrepModel.customer_name,
                      PrepModel.customer_surname,
                  ))
                  .all())

        candidates = []  # (sort_bucket, sort_remaining, order_date, order, item)
        for order in orders:
            if order.order_number in archived_numbers:
                continue
            details = _parse_details(order)
            # Sadece TEK ürünlü siparişler
            if len(details) != 1:
                continue
            item = details[0]
            if not _matches(item.get('barcode', '')):
                continue

            rs = _remaining_seconds(order.agreed_delivery_date)
            order_date = order.order_date or datetime.min
            if rs is not None and rs < threshold_seconds:
                # Kova 1 — acil: en az kalan (gecikmiş = en negatif) önce
                candidates.append((0, rs, order_date, order, item))
            else:
                # Kova 2 — standart: cut-off asc (tarihi yoksa +inf = en sona), sonra FIFO
                rs_key = rs if rs is not None else float('inf')
                candidates.append((1, rs_key, order_date, order, item))

        if not candidates:
            return jsonify({
                "found": False,
                "message": "Bu barkod için bekleyen tek ürünlü yeni sipariş yok.",
            }), 200

        candidates.sort(key=lambda c: (c[0], c[1], c[2]))
        _bucket, rs_sel, _od, order, item = candidates[0]

        # Ürün görüntü bilgileri (Product tablosundan, yoksa sipariş detayından)
        product = Product.query.filter_by(barcode=scanned).first()
        model_code = (product.product_main_id if product and product.product_main_id
                      else item.get('sku', 'N/A'))
        color = product.color if product and product.color else item.get('color', 'N/A')
        size = product.size if product and product.size else item.get('size', 'N/A')

        # Bilgi amaçlı raf önerisi (stoklu ilk raf) — picker zaten ürünü elinde tutuyor.
        raf = (RafUrun.query
               .filter(RafUrun.urun_barkodu == scanned, RafUrun.adet > 0)
               .order_by(RafUrun.raf_kodu.asc())
               .first())

        return jsonify({
            "found": True,
            "order_number": order.order_number,
            "redirect_url": url_for('siparis_hazirla.index',
                                    order_number=order.order_number, manuel=1),
            "model_code": model_code,
            "color": color,
            "size": size,
            "barcode": scanned,
            "raf": raf.raf_kodu if raf else None,
            "customer_name": ' '.join(filter(None, [order.customer_name, order.customer_surname])),
            "remaining_time": _format_remaining(rs_sel if rs_sel != float('inf') else None),
            "urgent": _bucket == 0,
        }), 200

    except Exception as e:
        traceback.print_exc()
        return jsonify({"found": False, "message": f"Hata: {e}"}), 500
