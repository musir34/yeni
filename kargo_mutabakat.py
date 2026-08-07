# kargo_mutabakat.py — 🚚 Kargo Fatura Mutabakatı (MNG)
# Kargo firmasının fatura Excel'ini sistemdeki kargolarla karşılaştırır:
#  - Birebir eşleşme (takip no / müşteri ref no)
#  - Bulanık eşleşme (alıcı adı + tarih yakınlığı) → "Şüpheli" listesi
#  - Desi bazlı tarife fiyat kontrolü
from flask import Blueprint, request, render_template, redirect, url_for, flash, send_file
import pandas as pd
import os
import json
import logging
from io import BytesIO
from datetime import datetime, timedelta
from difflib import SequenceMatcher
from werkzeug.utils import secure_filename

from user_logs import log_user_action
from login_logout import login_required, roles_required

logger = logging.getLogger(__name__)

try:
    from .models import (
        db, OrderCreated, OrderHazirlaniyor, OrderPicking, OrderShipped, OrderDelivered,
        OrderCancelled, OrderArchived, OrderReadyToShip, Archive, ReturnOrder,
        Degisim, YeniSiparis, Order, UretimSiparis,
    )
except ImportError:
    from models import (
        db, OrderCreated, OrderHazirlaniyor, OrderPicking, OrderShipped, OrderDelivered,
        OrderCancelled, OrderArchived, OrderReadyToShip, Archive, ReturnOrder,
        Degisim, YeniSiparis, Order, UretimSiparis,
    )

kargo_mutabakat_bp = Blueprint('kargo_mutabakat_bp', __name__)

TARIFE_FILE = './uploads/kargo_tarife.json'
MUTABAKAT_UPLOAD_DIR = './uploads/kargo_mutabakat'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
os.makedirs(MUTABAKAT_UPLOAD_DIR, exist_ok=True)

ORDER_TABLES = [
    OrderCreated, OrderHazirlaniyor, OrderPicking, OrderShipped,
    OrderDelivered, OrderCancelled, OrderArchived, OrderReadyToShip,
]
TABLE_LABELS = {
    'orders_created': 'Yeni', 'orders_hazirlaniyor': 'Hazırlanıyor',
    'orders_picking': 'İşleme Alınan', 'orders_shipped': 'Kargoda',
    'orders_delivered': 'Teslim', 'orders_cancelled': 'İptal',
    'orders_archived': 'Arşiv (statü)', 'orders_ready_to_ship': 'Hazır Gönderim',
}


def _platform_label(source, status_label):
    """source kolonundan platform etiketi: 'Trendyol Siparişi (Kargoda)' gibi."""
    src = (source or 'trendyol').strip().lower()
    if src == 'shopify':
        platform = 'Shopify Siparişi'
    elif src == 'woocommerce':
        platform = 'WooCommerce Siparişi'
    else:
        platform = 'Trendyol Siparişi'
    return f"{platform} ({status_label})" if status_label else platform

# Bulanık eşleşme eşikleri (ayar için tek yer)
NAME_SIMILARITY_THRESHOLD = 0.85
DATE_WINDOW_DAYS = 15
CANDIDATE_LOOKBACK_DAYS = 45   # fatura aralığından ne kadar geriye aday taransın
PRICE_TOLERANCE_TL = 0.05


# ──────────────────────────── Yardımcılar ────────────────────────────

def _tr_upper(s):
    """Türkçe'ye uygun büyük harf + boşluk normalizasyonu (i→İ, ı→I)."""
    if not s:
        return ''
    s = str(s).replace('i', 'İ').replace('ı', 'I').upper()
    return ' '.join(s.split())


def _parse_tr_number(val):
    """'117,98' / '1.234,56' / '144.9' → float. Boş/NaN → None."""
    s = str(val if val is not None else '').strip()
    if not s or s.lower() in ('nan', 'none'):
        return None
    s = s.replace(' ', '')
    if ',' in s:
        s = s.replace('.', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return None


def _parse_tr_date(val):
    """Excel tarihini datetime'a çevirir (dd.mm.yyyy öncelikli)."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, (pd.Timestamp, datetime)):
        return val.to_pydatetime() if isinstance(val, pd.Timestamp) else val
    s = str(val).strip()
    if not s or s.lower() == 'nan':
        return None
    for fmt in ("%d.%m.%Y %H:%M", "%d.%m.%Y", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None


def _clean_tracking(val):
    """GONDERI NO'yu normalize eder: yalnızca rakamlar."""
    s = str(val if val is not None else '').strip()
    if not s or s.lower() == 'nan':
        return ''
    return ''.join(ch for ch in s if ch.isdigit())


def _load_tarife():
    if os.path.exists(TARIFE_FILE):
        with open(TARIFE_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []


def _save_tarife(bands):
    with open(TARIFE_FILE, 'w', encoding='utf-8') as f:
        json.dump(bands, f, ensure_ascii=False, indent=2)


def _read_invoice(path):
    """MNG fatura Excel'ini okur. Satır 0 tarih bandı → header=1.
    dtype=str ZORUNLU: GONDERI NO float'a dönüp bozulmasın."""
    ext = path.rsplit('.', 1)[-1].lower()
    df = pd.read_excel(path, header=1, dtype=str, engine='openpyxl' if ext == 'xlsx' else None)
    df.columns = [str(c).strip() for c in df.columns]
    if 'GONDERI NO' not in df.columns:
        raise ValueError("Excel'de 'GONDERI NO' sütunu bulunamadı — MNG detaylı fatura listesi bekleniyor")
    return df


def _cell(row, key):
    """Hücreyi temiz metin olarak döner (NaN/boş → '')."""
    v = row.get(key)
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return ''
    s = str(v).strip()
    return '' if s.lower() == 'nan' else s


def _row_to_dict(row):
    """Fatura satırını sonuç tablolarında gösterilecek sözlüğe çevirir."""
    return {
        'sira_no': _cell(row, 'SIRA NO'),
        'fatura_no': _cell(row, 'FATURA NO'),
        'fatura_tarihi': _cell(row, 'FATURA KESIM TARIHI'),
        'gonderi_no': _clean_tracking(row.get('GONDERI NO')),
        'alici': _cell(row, 'ALICI MUSTERI'),
        'alici_il': _cell(row, 'ALICI IL'),
        'ref_no': _cell(row, 'MUSTERI REF NO'),
        'desi': _parse_tr_number(row.get('KG DESI')),
        'tasima_ucreti': _parse_tr_number(row.get('TASIMA UCRETI')),
        'teslim_ucreti': _parse_tr_number(row.get('ADRESE TESLIM UCRETI')),
        'alim_ucreti': _parse_tr_number(row.get('ADRESTEN ALIM UCRETI')),
        'genel_toplam': _parse_tr_number(row.get('GENEL TOPLAM')),
        'teslim_tarihi': _cell(row, 'TESLIM TARIHI'),
        'teslim_alan': _cell(row, 'TESLIM ALAN'),
        'son_durum': _cell(row, 'KARGO SON DURUM ACIKLAMA'),
    }


# ──────────────────────────── Eşleştirme ────────────────────────────

def _exact_match_maps(tracking_set, ref_set):
    """Birebir eşleşme haritaları döner:
    tracking → (kaynak, sipariş no) ve ref_no → (kaynak, sipariş no).
    Not: orders_archived'a tam entity select yasak → her yerde with_entities."""
    tracking_map = {}
    ref_map = {}

    for cls in ORDER_TABLES:
        durum = TABLE_LABELS.get(cls.__tablename__, cls.__tablename__)
        if tracking_set:
            rows = (cls.query
                    .with_entities(cls.cargo_tracking_number, cls.order_number, cls.source)
                    .filter(cls.cargo_tracking_number.in_(tracking_set)).all())
            for tr, on, src in rows:
                tracking_map.setdefault(tr, (_platform_label(src, durum), on))
        if ref_set:
            rows = (cls.query
                    .with_entities(cls.order_number, cls.package_number, cls.source)
                    .filter((cls.order_number.in_(ref_set)) | (cls.package_number.in_(ref_set)))
                    .all())
            for on, pn, src in rows:
                if on in ref_set:
                    ref_map.setdefault(on, (_platform_label(src, durum), on))
                if pn in ref_set:
                    ref_map.setdefault(pn, (_platform_label(src, durum), on))

    if tracking_set:
        rows = (Archive.query
                .with_entities(Archive.shipping_barcode, Archive.order_number, Archive.source)
                .filter(Archive.shipping_barcode.in_(tracking_set)).all())
        for tr, on, src in rows:
            tracking_map.setdefault(tr, (_platform_label(src, 'Arşiv'), on))

        rows = (ReturnOrder.query
                .with_entities(ReturnOrder.cargo_tracking_number, ReturnOrder.order_number)
                .filter(ReturnOrder.cargo_tracking_number.in_(tracking_set)).all())
        for tr, on in rows:
            tracking_map.setdefault(tr, ('İade', on))

        rows = (Degisim.query
                .with_entities(Degisim.musteri_kargo_takip, Degisim.kargo_kodu, Degisim.degisim_no)
                .filter((Degisim.musteri_kargo_takip.in_(tracking_set)) | (Degisim.kargo_kodu.in_(tracking_set)))
                .all())
        for mkt, kk, dn in rows:
            if mkt in tracking_set:
                tracking_map.setdefault(mkt, ('Değişim (gelen)', dn))
            if kk in tracking_set:
                tracking_map.setdefault(kk, ('Değişim', dn))

        # Eski ana 'orders' tablosu (geriye dönük uyum — iki ayrı takip kolonu var)
        rows = (Order.query
                .with_entities(Order.cargo_tracking_number, Order.shipping_barcode, Order.order_number)
                .filter((Order.cargo_tracking_number.in_(tracking_set)) | (Order.shipping_barcode.in_(tracking_set)))
                .all())
        for tr, sb, on in rows:
            if tr in tracking_set:
                tracking_map.setdefault(tr, ('Sipariş (Eski Tablo)', on))
            if sb in tracking_set:
                tracking_map.setdefault(sb, ('Sipariş (Eski Tablo)', on))

    if ref_set:
        rows = (Archive.query
                .with_entities(Archive.order_number, Archive.package_number, Archive.source)
                .filter((Archive.order_number.in_(ref_set)) | (Archive.package_number.in_(ref_set)))
                .all())
        for on, pn, src in rows:
            if on in ref_set:
                ref_map.setdefault(on, (_platform_label(src, 'Arşiv'), on))
            if pn in ref_set:
                ref_map.setdefault(pn, (_platform_label(src, 'Arşiv'), on))

    return tracking_map, ref_map


def _fuzzy_candidates(window_start, window_end):
    """Ad+tarih bulanık eşleşmesi için aday havuzu: (normalize ad, tarih, kaynak, sipariş no)."""
    candidates = []

    for cls in ORDER_TABLES:
        durum = TABLE_LABELS.get(cls.__tablename__, cls.__tablename__)
        rows = (cls.query
                .with_entities(cls.customer_name, cls.customer_surname, cls.order_date,
                               cls.order_number, cls.source)
                .filter(cls.order_date >= window_start, cls.order_date <= window_end).all())
        for ad, soyad, tarih, on, src in rows:
            candidates.append((_tr_upper(f"{ad or ''} {soyad or ''}"), tarih, _platform_label(src, durum), on))

    rows = (Archive.query
            .with_entities(Archive.customer_name, Archive.customer_surname, Archive.order_date,
                           Archive.order_number, Archive.source)
            .filter(Archive.order_date >= window_start, Archive.order_date <= window_end).all())
    for ad, soyad, tarih, on, src in rows:
        candidates.append((_tr_upper(f"{ad or ''} {soyad or ''}"), tarih, _platform_label(src, 'Arşiv'), on))

    rows = (Degisim.query
            .with_entities(Degisim.ad, Degisim.soyad, Degisim.degisim_tarihi, Degisim.degisim_no)
            .filter(Degisim.degisim_tarihi >= window_start, Degisim.degisim_tarihi <= window_end).all())
    for ad, soyad, tarih, dn in rows:
        candidates.append((_tr_upper(f"{ad or ''} {soyad or ''}"), tarih, 'Değişim', dn))

    rows = (YeniSiparis.query
            .with_entities(YeniSiparis.musteri_adi, YeniSiparis.musteri_soyadi,
                           YeniSiparis.siparis_tarihi, YeniSiparis.siparis_no)
            .filter(YeniSiparis.siparis_tarihi >= window_start, YeniSiparis.siparis_tarihi <= window_end).all())
    for ad, soyad, tarih, sn in rows:
        candidates.append((_tr_upper(f"{ad or ''} {soyad or ''}"), tarih, 'Yeni Sipariş (Manuel)', sn))

    # Eski ana 'orders' tablosu
    rows = (Order.query
            .with_entities(Order.customer_name, Order.customer_surname, Order.order_date, Order.order_number)
            .filter(Order.order_date >= window_start, Order.order_date <= window_end).all())
    for ad, soyad, tarih, on in rows:
        candidates.append((_tr_upper(f"{ad or ''} {soyad or ''}"), tarih, 'Sipariş (Eski Tablo)', on))

    # Üretim siparişleri (yaşam döngüsü tablolarının kopyası olsa da tam kapsam için)
    rows = (UretimSiparis.query
            .with_entities(UretimSiparis.customer_name, UretimSiparis.order_date, UretimSiparis.order_number)
            .filter(UretimSiparis.order_date >= window_start, UretimSiparis.order_date <= window_end).all())
    for ad, tarih, on in rows:
        candidates.append((_tr_upper(ad or ''), tarih, 'Üretim Siparişi', on))

    # shopify_orders / woo_orders tabloları (model sınıfı yok → ham SQL; hata olursa atla)
    from sqlalchemy import text
    for tablo, etiket in (('shopify_orders', 'Shopify Siparişi (Tablo)'),
                          ('woo_orders', 'WooCommerce Siparişi (Tablo)')):
        try:
            rows = db.session.execute(
                text(f"SELECT customer_first_name, customer_last_name, customer_name, "
                     f"date_created, order_number FROM {tablo} "
                     f"WHERE date_created >= :s AND date_created <= :e"),
                {'s': window_start, 'e': window_end}).fetchall()
            for ad, soyad, tam_ad, tarih, on in rows:
                isim = f"{ad or ''} {soyad or ''}".strip() or (tam_ad or '')
                candidates.append((_tr_upper(isim), tarih, etiket, on))
        except Exception as e:
            db.session.rollback()
            logger.warning(f"Kargo mutabakat {tablo} taraması atlandı: {e}")

    return [c for c in candidates if c[0]]


def _shopify_live_sources(window_start):
    """Canlı Shopify siparişlerini API'den tarar (arşivlenmemiş siparişler DB'de YOK).
    (tracking_map_ek, aday_listesi_ek) döner; API hatasında boş döner, analiz DB ile sürer."""
    try:
        from shopify_site.shopify_service import shopify_service
        if not shopify_service.is_configured():
            return {}, []
        # NOT: customer alanı read_customers scope ister (token'da YOK) — kullanma!
        # MNG'deki ALICI MUSTERI zaten kargo adresi adıdır → shippingAddress.name yeterli.
        q = """
        query KargoMutabakat($first: Int!, $query: String, $after: String) {
          orders(first: $first, query: $query, after: $after, sortKey: PROCESSED_AT, reverse: true) {
            pageInfo { hasNextPage endCursor }
            edges { node {
              name
              createdAt
              shippingAddress { name }
              fulfillments { trackingInfo { number } }
            } }
          }
        }
        """
        qf = f"created_at:>={window_start.strftime('%Y-%m-%d')}"
        tracking_map, candidates = {}, []
        after = None
        for _ in range(20):  # sayfa emniyeti (20×100 sipariş)
            res = shopify_service.run_graphql(q, {'first': 100, 'query': qf, 'after': after})
            if not res.get('success'):
                logger.warning(f"Kargo mutabakat Shopify sorgusu başarısız: {res.get('error')}")
                break
            data = (res.get('data') or {}).get('orders') or {}
            for e in data.get('edges', []):
                n = e.get('node') or {}
                order_name = n.get('name')
                full = (n.get('shippingAddress') or {}).get('name') or ''
                tarih = None
                created = str(n.get('createdAt') or '')[:10]
                if created:
                    try:
                        tarih = datetime.strptime(created, '%Y-%m-%d')
                    except ValueError:
                        pass
                for ff in (n.get('fulfillments') or []):
                    for ti in (ff.get('trackingInfo') or []):
                        num = _clean_tracking(ti.get('number'))
                        if num:
                            tracking_map.setdefault(num, ('Shopify Siparişi (Canlı)', order_name))
                if full:
                    candidates.append((_tr_upper(full), tarih, 'Shopify Siparişi (Canlı)', order_name))
            if not data.get('pageInfo', {}).get('hasNextPage'):
                break
            after = data.get('pageInfo', {}).get('endCursor')
        return tracking_map, candidates
    except Exception as e:
        logger.warning(f"Kargo mutabakat Shopify canlı tarama atlandı: {e}")
        return {}, []


def _best_fuzzy_match(norm_name, invoice_date, candidates, exact_name_index):
    """En iyi aday: (aday adı, benzerlik, kaynak, sipariş no) veya None.
    Önce ucuz birebir-ad eşitliği, sonra SequenceMatcher."""

    def _date_ok(cand_date):
        if invoice_date is None or cand_date is None:
            return True  # tarih bilinmiyorsa ada göre karar insanın
        return abs((invoice_date - cand_date).days) <= DATE_WINDOW_DAYS

    for cand in exact_name_index.get(norm_name, []):
        if _date_ok(cand[1]):
            return (cand[0], 1.0, cand[2], cand[3])

    best = None
    sm = SequenceMatcher()
    sm.set_seq2(norm_name)
    for cand_name, cand_date, label, on in candidates:
        if abs(len(cand_name) - len(norm_name)) > 6:
            continue
        if not _date_ok(cand_date):
            continue
        sm.set_seq1(cand_name)
        if sm.real_quick_ratio() < NAME_SIMILARITY_THRESHOLD:
            continue
        if sm.quick_ratio() < NAME_SIMILARITY_THRESHOLD:
            continue
        ratio = sm.ratio()
        if ratio >= NAME_SIMILARITY_THRESHOLD and (best is None or ratio > best[1]):
            best = (cand_name, ratio, label, on)
    return best


def _check_tarife(desi, tasima_ucreti, bands):
    """Tarife kontrolü: (durum, beklenen) — durum: 'ok' | 'anomali' | 'tarife_disi' | None."""
    if desi is None or tasima_ucreti is None or not bands:
        return None, None
    for b in bands:
        try:
            if float(b['min_desi']) <= desi <= float(b['max_desi']):
                beklenen = float(b['fiyat'])
                if abs(tasima_ucreti - beklenen) > PRICE_TOLERANCE_TL:
                    return 'anomali', beklenen
                return 'ok', beklenen
        except (KeyError, TypeError, ValueError):
            continue
    return 'tarife_disi', None


def _analyze(df):
    """Fatura DataFrame'ini analiz eder; kova sözlüğü döner."""
    bands = _load_tarife()

    rows = []
    for _, r in df.iterrows():
        d = _row_to_dict(r)
        if not d['gonderi_no']:
            continue
        d['_tarih'] = _parse_tr_date(d['fatura_tarihi'])
        rows.append(d)

    tracking_set = {d['gonderi_no'] for d in rows}
    ref_set = {d['ref_no'] for d in rows if d['ref_no']}
    tracking_map, ref_map = _exact_match_maps(tracking_set, ref_set)

    dates = [d['_tarih'] for d in rows if d['_tarih']]
    if dates:
        window_start = min(dates) - timedelta(days=CANDIDATE_LOOKBACK_DAYS)
        window_end = max(dates) + timedelta(days=1)
    else:
        window_end = datetime.utcnow()
        window_start = window_end - timedelta(days=90)

    # Canlı Shopify siparişleri (arşivlenmemişler DB'de yok — API'den taranır)
    shopify_tracking, shopify_cands = _shopify_live_sources(window_start)
    for k, v in shopify_tracking.items():
        tracking_map.setdefault(k, v)

    candidates = _fuzzy_candidates(window_start, window_end) + shopify_cands
    exact_name_index = {}
    for c in candidates:
        exact_name_index.setdefault(c[0], []).append(c)

    eslesen, supheli, bizde_yok, fiyat_anomali = [], [], [], []

    # 1. Adım: bizde var mı? Birebir eşleşenler sonraki işleme, olmayanlar 2. adıma
    for d in rows:
        if d['gonderi_no'] in tracking_map:
            label, on = tracking_map[d['gonderi_no']]
            d.update(match_type='takip_no', kaynak=label, siparis_no=on)
            eslesen.append(d)
            continue
        if d['ref_no'] and d['ref_no'] in ref_map:
            label, on = ref_map[d['ref_no']]
            d.update(match_type='ref_no', kaynak=label, siparis_no=on)
            eslesen.append(d)
            continue

        # 2. Adım: şüpheli mi? (ad + tarih benzerliği) Değilse → Bizde Yok
        best = _best_fuzzy_match(_tr_upper(d['alici']), d['_tarih'], candidates, exact_name_index)
        if best:
            d.update(match_type='ad_tarih', aday_ad=best[0], benzerlik=round(best[1] * 100),
                     kaynak=best[2], siparis_no=best[3])
            supheli.append(d)
        else:
            bizde_yok.append(d)

    # 3. Adım: fiyat kontrolü — yalnızca bizim olan (eşleşen + şüpheli) kargolarda.
    # "Bizde Yok" satırları zaten komple itiraz konusu, fiyatına ayrıca bakılmaz.
    for d in eslesen + supheli:
        durum, beklenen = _check_tarife(d['desi'], d['tasima_ucreti'], bands)
        d['tarife_durum'] = durum
        d['beklenen_fiyat'] = beklenen
        if durum in ('anomali', 'tarife_disi'):
            fiyat_anomali.append(d)

    for d in rows:
        d.pop('_tarih', None)

    return {
        'toplam': len(rows),
        'eslesen': eslesen,
        'supheli': supheli,
        'bizde_yok': bizde_yok,
        'fiyat_anomali': fiyat_anomali,
        'itiraz_tutari': round(sum(d['genel_toplam'] or 0 for d in bizde_yok), 2),
        'shopify_canli': len(shopify_cands),
        'aday_sayisi': len(candidates),
    }


# ──────────────────────────── Route'lar ────────────────────────────

@kargo_mutabakat_bp.route('/kargo-mutabakat', methods=['GET'])
@login_required
@roles_required('admin')
def kargo_mutabakat_page():
    return render_template('kargo_mutabakat.html', tarife=_load_tarife(), sonuc=None, saved_filename=None)


@kargo_mutabakat_bp.route('/kargo-mutabakat/analiz', methods=['POST'])
@login_required
@roles_required('admin')
def kargo_mutabakat_analiz():
    f = request.files.get('excel_file')
    if not f or not f.filename:
        flash('Dosya yüklenmedi', 'danger')
        return redirect(url_for('kargo_mutabakat_bp.kargo_mutabakat_page'))

    filename = secure_filename(f.filename)
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
    if ext not in ALLOWED_EXTENSIONS:
        flash('Geçersiz dosya uzantısı (xls/xlsx bekleniyor)', 'danger')
        return redirect(url_for('kargo_mutabakat_bp.kargo_mutabakat_page'))

    saved_filename = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{filename}"
    save_path = os.path.join(MUTABAKAT_UPLOAD_DIR, saved_filename)
    f.save(save_path)

    try:
        df = _read_invoice(save_path)
        sonuc = _analyze(df)
    except Exception as e:
        logger.error(f"Kargo mutabakat analiz hatası ({saved_filename}): {e}", exc_info=True)
        flash(f'Analiz hatası: {e}', 'danger')
        return redirect(url_for('kargo_mutabakat_bp.kargo_mutabakat_page'))

    try:
        log_user_action("UPLOAD", {
            "işlem_açıklaması": (
                f"Kargo mutabakat analizi — {filename}: {sonuc['toplam']} satır, "
                f"{len(sonuc['eslesen'])} eşleşti, {len(sonuc['supheli'])} şüpheli, "
                f"{len(sonuc['bizde_yok'])} bizde yok"
            ),
            "sayfa": "Kargo Mutabakat",
        })
    except Exception:
        pass

    return render_template('kargo_mutabakat.html', tarife=_load_tarife(),
                           sonuc=sonuc, saved_filename=saved_filename)


@kargo_mutabakat_bp.route('/kargo-mutabakat/tarife', methods=['POST'])
@login_required
@roles_required('admin')
def kargo_mutabakat_tarife():
    min_list = request.form.getlist('min_desi')
    max_list = request.form.getlist('max_desi')
    fiyat_list = request.form.getlist('fiyat')
    bands = []
    for mn, mx, fy in zip(min_list, max_list, fiyat_list):
        mn_v, mx_v, fy_v = _parse_tr_number(mn), _parse_tr_number(mx), _parse_tr_number(fy)
        if mn_v is None or mx_v is None or fy_v is None:
            continue
        bands.append({'min_desi': mn_v, 'max_desi': mx_v, 'fiyat': fy_v})
    bands.sort(key=lambda b: b['min_desi'])
    _save_tarife(bands)
    flash(f'Tarife kaydedildi ({len(bands)} bant)', 'success')
    return redirect(url_for('kargo_mutabakat_bp.kargo_mutabakat_page'))


@kargo_mutabakat_bp.route('/kargo-mutabakat/itiraz-excel/<filename>', methods=['GET'])
@login_required
@roles_required('admin')
def kargo_mutabakat_itiraz_excel(filename):
    # Path traversal koruması: sadece MUTABAKAT_UPLOAD_DIR içindeki kayıtlı dosyalar
    safe_name = secure_filename(filename)
    save_path = os.path.join(MUTABAKAT_UPLOAD_DIR, safe_name)
    if safe_name != filename or not os.path.isfile(save_path):
        flash('Dosya bulunamadı — lütfen Excel\'i yeniden yükleyin', 'danger')
        return redirect(url_for('kargo_mutabakat_bp.kargo_mutabakat_page'))

    try:
        df = _read_invoice(save_path)
        sonuc = _analyze(df)
    except Exception as e:
        logger.error(f"Kargo mutabakat itiraz-excel hatası ({safe_name}): {e}", exc_info=True)
        flash(f'İtiraz listesi oluşturulamadı: {e}', 'danger')
        return redirect(url_for('kargo_mutabakat_bp.kargo_mutabakat_page'))

    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    bold = Font(bold=True)

    def _fill_sheet(ws, items, extra_headers=(), extra_getter=None):
        headers = ['Sıra No', 'Fatura No', 'Fatura Tarihi', 'Gönderi No', 'Alıcı', 'Alıcı İl',
                   'Desi', 'Taşıma Ücreti', 'Genel Toplam', 'Teslim Tarihi', 'Teslim Alan',
                   'Son Durum'] + list(extra_headers)
        ws.append(headers)
        for c in ws[1]:
            c.font = bold
        for d in items:
            row = [d['sira_no'], d['fatura_no'], d['fatura_tarihi'], d['gonderi_no'],
                   d['alici'], d['alici_il'], d['desi'], d['tasima_ucreti'],
                   d['genel_toplam'], d['teslim_tarihi'], d['teslim_alan'], d['son_durum']]
            if extra_getter:
                row += extra_getter(d)
            ws.append(row)

    ws1 = wb.active
    ws1.title = 'Bizde Yok'
    _fill_sheet(ws1, sonuc['bizde_yok'])

    ws2 = wb.create_sheet('Şüpheli')
    _fill_sheet(ws2, sonuc['supheli'],
                extra_headers=('Benzer Müşteri', 'Benzerlik %', 'Kaynak', 'Sipariş No'),
                extra_getter=lambda d: [d.get('aday_ad'), d.get('benzerlik'), d.get('kaynak'), d.get('siparis_no')])

    ws3 = wb.create_sheet('Fiyat Anomalileri')
    _fill_sheet(ws3, sonuc['fiyat_anomali'],
                extra_headers=('Beklenen Fiyat', 'Fark', 'Durum', 'Kaynak', 'Sipariş No'),
                extra_getter=lambda d: [
                    d.get('beklenen_fiyat'),
                    round((d['tasima_ucreti'] or 0) - (d.get('beklenen_fiyat') or 0), 2) if d.get('beklenen_fiyat') is not None else None,
                    'Tarife dışı desi' if d.get('tarife_durum') == 'tarife_disi' else 'Fiyat farkı',
                    d.get('kaynak') or 'Bizde Yok',
                    d.get('siparis_no'),
                ])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"itiraz_listesi_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
